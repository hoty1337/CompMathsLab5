import sys
import numpy as np
import pandas as pd
import openpyxl as opx
import matplotlib.pyplot as plt
from scipy.interpolate import lagrange, BarycentricInterpolator

# Исходные данные в виде таблицы x, y
x = np.array([1.25, 1.5, 1.75, 2.0])
y = np.array([1.2438, 1.3757, 1.5628, 1.8031])


# Функция для вывода таблицы конечных разностей
def finite_differences_table(x, y):
	n = len(y)
	F = np.zeros((n, n))
	F[:, 0] = y
	for j in range(1, n):
		for i in range(n - j):
			F[i][j] = (F[i + 1][j - 1] - F[i][j - 1]) / (x[i + j] - x[i])
	return F


# Функция для интерполяции многочленом Лагранжа
def lagrange_interpolation(x, y, x_interp):
	poly = lagrange(x, y)
	y_interp = poly(x_interp)
	return y_interp


# Функция для интерполяции многочленом Ньютона с конечными разностями
def newton_interpolation(x, y, x_interp):
	poly = BarycentricInterpolator(x, y)
	y_interp = poly(x_interp)
	return y_interp


# Функция для интерполяции многочленом Гаусса
def gauss_interpolation(x, y, x_interp):
	interp = BarycentricInterpolator(x, y)
	y_interp = interp(x_interp)
	return y_interp


try:
	check = (input(
		"Если вы хотите воспользоваться готовым датасетом выберите его номер (1,2 или 3), иначе введите пустую строку: "))
	if check == "1":
		data = np.loadtxt(r'dataset/data.txt', delimiter=',', skiprows=1)
		x = data[:, 0]
		y = data[:, 1]
	elif check == "2":
		workbook = opx.load_workbook(r'dataset/data.xlsx')
		worksheet = workbook.active
		worksheet = workbook['Лист1']
		x = [i[0] for i in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True)
		     if i[0] is not None][1:]
		y = [i[0] for i in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=2, max_col=2, values_only=True)
		     if i[0] is not None][1:]
	elif check == "3":
		data = pd.read_csv(r'dataset/data.csv')
		x = data.iloc[:, 0].values
		y = data.iloc[:, 1].values
	else:
		try:
			print("Введите значения x через запятую:")
			x = np.array(input().split(',')).astype(float)
		except Exception as e:
			print("Ошибка ввода x:", e)
			sys.exit()

		try:
			print("Введите значения y через запятую:")
			y = np.array(input().split(',')).astype(float)
		except Exception as e:
			print("Ошибка ввода y:", e)
			sys.exit()
except Exception as e:
	print("Ошибка выбора данных:", e)
	sys.exit()

try:
	x_interp = float(input("Введите значение x для интерполяции: "))
except Exception as e:
	print("Ошибка ввода y:", e)
	sys.exit()

try:
	s = input("Чтобы выбрать sin(x) введите s. Чтобы выбрать cos(x) введите c. Чтобы выбрать tan(x) введите t: ")
except Exception as e:
	print("Ошибка выбора функции:", e)
	sys.exit()

try:
	start_interval = float(
		input("Введите начальное значение для интервала интерполяции (значение вводится в pi): ")) * np.pi
	finish_interval = float(
		input("Введите конечное значение для интервала интерполяции (значение вводится в pi): ")) * np.pi
	step_interval = float(input("Введите шаг для интервала интерполяции: "))
	# point_count = int(input("Введите кол-во точек: "))
	if start_interval > finish_interval:
		print("Ошибка ввода интервала: значение начального интервала больше конечного")
		sys.exit()
except Exception as e:
	print("Ошибка ввода интервала:", e)
	sys.exit()

# Вывод таблицы конечных разностей
try:
	F = finite_differences_table(x, y)
	print("Таблица конечных разностей:")
	print(F)
except Exception as e:
	print("Ошибка построения таблицы:", e)

# Вычисление значений интерполяционных многочленов и сравнение результатов
try:
	y_lagrange = lagrange_interpolation(x, y, x_interp)
	y_newton = newton_interpolation(x, y, x_interp)
	y_gauss = gauss_interpolation(x, y, x_interp)
	print("Многочлен Лагранжа: ", y_lagrange)
	print("Многочлен Ньютона: ", y_newton)
	print("Многочлен Гаусса: ", y_gauss)
except Exception as e:
	print("Ошибка расчета многочленов:", e)


# Функция для заданной функции f(x)
def sin(x):
	return np.sin(x)


def cos(x):
	return np.cos(x)


def tan(x):
	return np.tan(x)


# Интерполяция функции f(x) на заданном отрезке
x = np.arange(start_interval, finish_interval + step_interval, step_interval)
if s == "s" or s == "S":
	print("Выбрана функция sin(x)")
	y = sin(x)
elif s == "c" or s == "C":
	print("Выбрана функция cos(x)")
	y = cos(x)
elif s == "t" or s == "T":
	print("Выбрана функция tan(x)")
	y = tan(x)
else:
	print("Вы не выбрали ни одной функции, базово выбрана функция sin(x)")
	y = sin(x)

# Построение графика
fig, ax = plt.subplots()
xx = np.linspace(min(x), max(x), round((max(x) - min(x)) / step_interval))
if s == "c" or s == "C":
	yy = cos(xx)
elif s == "t" or s == "T":
	yy = tan(xx)
else:
	yy = sin(xx)


ax.plot(xx, yy, label='Исходная функция')
ax.plot(x, y, 'o', label='Узлы интерполяции')
ax.plot(xx, newton_interpolation(x, y, xx), label='Многочлен Ньютона', alpha=0.5)
ax.plot(xx, gauss_interpolation(x, y, xx), label='Многочлен Гаусса', alpha=0.5)
ax.legend()
plt.show()
